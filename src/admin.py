import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import shutil
import os
from datetime import datetime

# Predefined office locations with coordinates
OFFICE_LOCATIONS = {
    "Kantor Gubernur Sulawesi Selatan": "-5.1508,119.4321",
    "Kantor Walikota Makassar": "-5.1531,119.4319", 
    "Kantor DPRD Makassar": "-5.1534,119.4326",
    "Universitas Hasanuddin": "-5.1308,119.4863",
    "Kantor BPN Makassar": "-5.1514,119.4321",
    "RS Wahidin Sudirohusodo": "-5.1397,119.4902"
}

# Predefined cities with coordinates
CITIES = {
    "Makassar": "-5.1477,119.4327",
    "Jakarta": "-6.2088,106.8456",
    "Surabaya": "-7.2575,112.7521", 
    "Bandung": "-6.9175,107.6191",
    "Medan": "3.5952,98.6722",
    "Semarang": "-6.9932,110.4203"
}

class AdminSystem:
    def __init__(self, db):
        print("\nInitializing Admin System...")
        self.db = db
        self.window = tk.Toplevel()
        self.window.title("Admin Panel - Sistem Absensi")
        self.window.geometry("1000x700")
        self.window.grab_set()  # Make window modal
        
        # Center window
        self.center_window()
        
        # Setup custom styles
        self.setup_styles()
        
        # Selected photo path
        self.selected_photo_path = None
        self.current_user_id = None  # For editing
        
        # Create notebook for tabs
        self.notebook = ttk.Notebook(self.window)
        self.notebook.pack(expand=True, fill='both', padx=10, pady=10)
        
        # Create tabs
        self.create_user_management_tab()
        self.create_attendance_report_tab()
        print("Admin System initialized")
        
    def center_window(self):
        """Center window on screen"""
        self.window.update_idletasks()
        width = self.window.winfo_width()
        height = self.window.winfo_height()
        x = (self.window.winfo_screenwidth() // 2) - (width // 2)
        y = (self.window.winfo_screenheight() // 2) - (height // 2)
        self.window.geometry(f'{width}x{height}+{x}+{y}')
        
    def setup_styles(self):
        """Setup custom ttk styles"""
        style = ttk.Style()
        
        # Configure main styles
        style.configure(
            "Admin.TFrame",
            background="#f5f6fa"
        )
        style.configure(
            "Card.TFrame",
            background="white",
            relief="solid",
            borderwidth=1
        )
        style.configure(
            "AdminTitle.TLabel",
            font=('Helvetica', 16, 'bold'),
            background="#f5f6fa",
            padding=10
        )
        style.configure(
            "Header.TLabel",
            font=('Helvetica', 12, 'bold'),
            background="white",
            padding=5
        )
        style.configure(
            "Action.TButton",
            padding=8,
            font=('Helvetica', 10)
        )
        style.configure(
            "Primary.TButton",
            padding=8,
            font=('Helvetica', 10, 'bold')
        )
        style.configure(
            "Danger.TButton",
            padding=8,
            font=('Helvetica', 10),
            background="#e74c3c"
        )
        
        # Configure Treeview
        style.configure(
            "Treeview",
            background="white",
            fieldbackground="white",
            rowheight=25
        )
        style.configure(
            "Treeview.Heading",
            font=('Helvetica', 10, 'bold'),
            padding=5
        )
        
        # Map hover effects
        style.map(
            "Treeview",
            background=[('selected', '#3498db')],
            foreground=[('selected', 'white')]
        )

    def create_user_management_tab(self):
        """Create user management interface with improved UI"""
        tab = ttk.Frame(self.notebook, style="Admin.TFrame")
        self.notebook.add(tab, text="Manajemen User")
        
        # Add User Frame
        add_frame = ttk.Frame(tab, style="Card.TFrame")
        add_frame.pack(pady=10, padx=20, fill='x')
        
        # Title
        ttk.Label(
            add_frame,
            text="Tambah/Edit User",
            style="Header.TLabel"
        ).pack(pady=(10,5), padx=10)
        
        # Form container for 2 columns
        form_frame = ttk.Frame(add_frame)
        form_frame.pack(padx=20, pady=10, fill='x')
        form_frame.columnconfigure(1, weight=1)
        form_frame.columnconfigure(3, weight=1)
        
        # Left column
        # Name
        ttk.Label(
            form_frame,
            text="Nama:",
            style="Header.TLabel"
        ).grid(row=0, column=0, padx=5, pady=5, sticky='e')
        
        self.name_entry = ttk.Entry(form_frame, width=30)
        self.name_entry.grid(row=0, column=1, padx=5, pady=5, sticky='ew')
        
        # Photo
        ttk.Label(
            form_frame,
            text="Foto:",
            style="Header.TLabel"
        ).grid(row=1, column=0, padx=5, pady=5, sticky='e')
        
        photo_frame = ttk.Frame(form_frame)
        photo_frame.grid(row=1, column=1, padx=5, pady=5, sticky='ew')
        
        self.photo_entry = ttk.Entry(photo_frame, width=25, state='readonly')
        self.photo_entry.pack(side=tk.LEFT, padx=(0,5))
        
        self.photo_button = ttk.Button(
            photo_frame,
            text="Pilih Foto",
            command=self.select_photo,
            style="Action.TButton"
        )
        self.photo_button.pack(side=tk.LEFT)
        
        # Right column
        # Home Location
        ttk.Label(
            form_frame,
            text="Kota (WFH):",
            style="Header.TLabel"
        ).grid(row=0, column=2, padx=5, pady=5, sticky='e')
        
        self.home_city = ttk.Combobox(
            form_frame,
            width=28,
            values=sorted(list(CITIES.keys())),
            state='readonly'
        )
        self.home_city.grid(row=0, column=3, padx=5, pady=5, sticky='ew')
        
        # Office Location
        ttk.Label(
            form_frame,
            text="Kantor (WFO):",
            style="Header.TLabel"
        ).grid(row=1, column=2, padx=5, pady=5, sticky='e')
        
        self.office_loc = ttk.Combobox(
            form_frame,
            width=28,
            values=sorted(list(OFFICE_LOCATIONS.keys())),
            state='readonly'
        )
        self.office_loc.grid(row=1, column=3, padx=5, pady=5, sticky='ew')
        
        # Buttons frame
        button_frame = ttk.Frame(add_frame)
        button_frame.pack(pady=10)
        
        # Add/Update Button
        self.submit_button = ttk.Button(
            button_frame,
            text="Tambah User",
            command=self.submit_user,
            style="Primary.TButton"
        )
        self.submit_button.pack(side=tk.LEFT, padx=5)
        
        # Cancel Edit Button (hidden by default)
        self.cancel_button = ttk.Button(
            button_frame,
            text="Batal Edit",
            command=self.cancel_edit,
            style="Action.TButton"
        )

        # User List Frame
        list_frame = ttk.Frame(tab, style="Card.TFrame")
        list_frame.pack(pady=10, padx=20, fill='both', expand=True)
        
        # Header with action buttons
        header_frame = ttk.Frame(list_frame)
        header_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Label(
            header_frame,
            text="Daftar User",
            style="Header.TLabel"
        ).pack(side=tk.LEFT)
        
        # Action buttons on right
        action_frame = ttk.Frame(header_frame)
        action_frame.pack(side=tk.RIGHT)
        
        ttk.Button(
            action_frame,
            text="Edit",
            command=self.edit_user,
            style="Action.TButton"
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            action_frame,
            text="Hapus",
            command=self.delete_user,
            style="Danger.TButton"
        ).pack(side=tk.LEFT, padx=5)

        # Create treeview with container
        tree_container = ttk.Frame(list_frame)
        tree_container.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Create treeview
        columns = ("ID", "Nama", "Photo Path", "Tanggal Dibuat")
        self.user_tree = ttk.Treeview(
            tree_container,
            columns=columns,
            show='headings',
            selectmode='browse'
        )
        
        # Define columns
        self.user_tree.heading('ID', text='ID')
        self.user_tree.heading('Nama', text='Nama')
        self.user_tree.heading('Photo Path', text='Photo Path')
        self.user_tree.heading('Tanggal Dibuat', text='Tanggal Dibuat')
        
        self.user_tree.column('ID', width=50, anchor='center')
        self.user_tree.column('Nama', width=200)
        self.user_tree.column('Photo Path', width=250)
        self.user_tree.column('Tanggal Dibuat', width=150, anchor='center')
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(tree_container, orient="vertical", command=self.user_tree.yview)
        self.user_tree.configure(yscrollcommand=scrollbar.set)
        
        # Pack tree and scrollbar
        self.user_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Load initial data
        self.load_users()

    def create_attendance_report_tab(self):
        """Create attendance report interface with improved UI"""
        tab = ttk.Frame(self.notebook, style="Admin.TFrame")
        self.notebook.add(tab, text="Laporan Absensi")
        
        # Header frame
        header_frame = ttk.Frame(tab, style="Card.TFrame")
        header_frame.pack(fill='x', padx=20, pady=10)
        
        # Title and export button in header
        title_frame = ttk.Frame(header_frame)
        title_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Label(
            title_frame,
            text="Laporan Kehadiran",
            style="Header.TLabel"
        ).pack(side=tk.LEFT)
        
        ttk.Button(
            title_frame,
            text="Export ke Excel",
            command=self.export_attendance,
            style="Primary.TButton"
        ).pack(side=tk.RIGHT)
        
        # Filter frame (for future use)
        filter_frame = ttk.Frame(header_frame)
        filter_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Label(
            filter_frame,
            text="Filter:",
            style="Header.TLabel"
        ).pack(side=tk.LEFT, padx=(0,10))
        
        # Placeholder for future filter controls
        # Will add date range, user filter, etc.
        
        # Main content frame
        content_frame = ttk.Frame(tab, style="Card.TFrame")
        content_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Create treeview container
        tree_container = ttk.Frame(content_frame)
        tree_container.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Create treeview
        columns = ("ID", "Nama", "Tanggal", "Jam Masuk", "Jam Keluar", "Mode", "Status", "Lokasi")
        self.attendance_tree = ttk.Treeview(
            tree_container,
            columns=columns,
            show='headings',
            selectmode='browse'
        )
        
        # Define columns
        self.attendance_tree.heading('ID', text='ID')
        self.attendance_tree.heading('Nama', text='Nama')
        self.attendance_tree.heading('Tanggal', text='Tanggal')
        self.attendance_tree.heading('Jam Masuk', text='Jam Masuk')
        self.attendance_tree.heading('Jam Keluar', text='Jam Keluar')
        self.attendance_tree.heading('Mode', text='Mode')
        self.attendance_tree.heading('Status', text='Status')
        self.attendance_tree.heading('Lokasi', text='Lokasi')
        
        # Set column widths and alignments
        self.attendance_tree.column('ID', width=50, anchor='center')
        self.attendance_tree.column('Nama', width=150)
        self.attendance_tree.column('Tanggal', width=100, anchor='center')
        self.attendance_tree.column('Jam Masuk', width=100, anchor='center')
        self.attendance_tree.column('Jam Keluar', width=100, anchor='center')
        self.attendance_tree.column('Mode', width=80, anchor='center')
        self.attendance_tree.column('Status', width=80, anchor='center')
        self.attendance_tree.column('Lokasi', width=200)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(tree_container, orient="vertical", command=self.attendance_tree.yview)
        self.attendance_tree.configure(yscrollcommand=scrollbar.set)
        
        # Pack tree and scrollbar
        self.attendance_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Summary frame at bottom
        summary_frame = ttk.Frame(content_frame)
        summary_frame.pack(fill='x', padx=10, pady=5)
        
        # Will add summary statistics here
        
        # Load initial data
        self.load_attendance()

    def show_loading(self, message="Loading..."):
        """Show loading overlay"""
        self.loading_frame = ttk.Frame(
            self.window,
            style="Card.TFrame"
        )
        self.loading_frame.place(
            relx=0.5,
            rely=0.5,
            anchor='center'
        )
        
        self.loading_label = ttk.Label(
            self.loading_frame,
            text=message,
            font=('Helvetica', 12),
            padding=20
        )
        self.loading_label.pack(pady=10)
        
        # Create progress bar
        self.progress = ttk.Progressbar(
            self.loading_frame,
            length=200,
            mode='indeterminate'
        )
        self.progress.pack(pady=5)
        self.progress.start()
        
        self.window.update()

    def hide_loading(self):
        """Hide loading overlay"""
        if hasattr(self, 'loading_frame'):
            self.progress.stop()
            self.loading_frame.destroy()

    def submit_user(self):
        """Add new user or update existing user with improved feedback"""
        try:
            self.show_loading("Menyimpan data user...")
            
            # Get form data
            name = self.name_entry.get().strip()
            home_city = self.home_city.get().strip()
            office = self.office_loc.get().strip()
            
            # Validate inputs with detailed feedback
            if not name:
                self.hide_loading()
                messagebox.showerror(
                    "Error",
                    "Nama user harus diisi!",
                    parent=self.window
                )
                self.name_entry.focus()
                return
                
            if not home_city:
                self.hide_loading()
                messagebox.showerror(
                    "Error",
                    "Kota (WFH) harus dipilih!",
                    parent=self.window
                )
                self.home_city.focus()
                return
                
            if not office:
                self.hide_loading()
                messagebox.showerror(
                    "Error",
                    "Lokasi kantor (WFO) harus dipilih!",
                    parent=self.window
                )
                self.office_loc.focus()
                return
                
            # Get coordinates
            home_coord = CITIES[home_city]
            office_coord = OFFICE_LOCATIONS[office]
            
            if self.current_user_id:  # Update existing user
                if self.selected_photo_path:  # If new photo selected
                    photo_name = self.process_photo(name)
                else:
                    photo_name = None
                    
                success = self.db.update_user(
                    self.current_user_id,
                    name,
                    photo_name,
                    home_coord,
                    office_coord
                )
                
                if success:
                    self.hide_loading()
                    messagebox.showinfo(
                        "Sukses",
                        f"Data user {name} berhasil diupdate!",
                        parent=self.window
                    )
                    self.cancel_edit()
                else:
                    raise Exception("Gagal update user")
                    
            else:  # Add new user
                if not self.selected_photo_path:
                    self.hide_loading()
                    messagebox.showerror(
                        "Error",
                        "Foto harus dipilih untuk user baru!",
                        parent=self.window
                    )
                    return
                    
                # Process photo
                photo_name = self.process_photo(name)
                
                # Save to database
                if self.db.add_user(
                    name=name,
                    photo_path=photo_name,
                    home_location=home_coord,
                    office_location=office_coord
                ):
                    self.hide_loading()
                    messagebox.showinfo(
                        "Sukses",
                        f"User baru {name} berhasil ditambahkan!",
                        parent=self.window
                    )
                    self.clear_form()
                else:
                    raise Exception("Gagal menambahkan user")
                    
            # Refresh user list
            self.load_users()
            
        except Exception as e:
            print(f"Error submitting user: {str(e)}")
            self.hide_loading()
            messagebox.showerror(
                "Error",
                f"Gagal menyimpan user: {str(e)}",
                parent=self.window
            )

    def select_photo(self):
        """Open file dialog to select user photo with preview"""
        try:
            filename = filedialog.askopenfilename(
                title="Pilih Foto User",
                filetypes=[
                    ("Image files", "*.jpg *.jpeg *.png")
                ],
                parent=self.window
            )
            
            if filename:
                # Validate file size
                file_size = os.path.getsize(filename) / (1024 * 1024)  # Convert to MB
                if file_size > 5:  # 5MB limit
                    messagebox.showerror(
                        "Error",
                        "Ukuran file terlalu besar! Maksimal 5MB.",
                        parent=self.window
                    )
                    return
                
                self.selected_photo_path = filename
                self.photo_entry.configure(state='normal')
                self.photo_entry.delete(0, tk.END)
                self.photo_entry.insert(0, filename)
                self.photo_entry.configure(state='readonly')
                
        except Exception as e:
            print(f"Error selecting photo: {str(e)}")
            messagebox.showerror(
                "Error",
                f"Gagal memilih foto: {str(e)}",
                parent=self.window
            )

    def edit_user(self):
        """Start editing selected user with improved feedback"""
        try:
            # Get selected item
            selection = self.user_tree.selection()
            if not selection:
                messagebox.showwarning(
                    "Peringatan",
                    "Silakan pilih user yang akan diedit!",
                    parent=self.window
                )
                return
                
            # Get user ID
            user_id = self.user_tree.item(selection[0])['values'][0]
            
            # Show loading
            self.show_loading("Memuat data user...")
            
            # Get user data
            user = self.db.get_user_by_id(user_id)
            if not user:
                self.hide_loading()
                messagebox.showerror(
                    "Error",
                    "Data user tidak ditemukan!",
                    parent=self.window
                )
                return
                
            # Set form to edit mode
            self.current_user_id = user_id
            self.submit_button.configure(text="Update User")
            self.cancel_button.pack(side=tk.LEFT, padx=5)
            
            # Fill form with user data
            self.name_entry.delete(0, tk.END)
            self.name_entry.insert(0, user['name'])
            
            # Clear photo selection
            self.selected_photo_path = None
            self.photo_entry.configure(state='normal')
            self.photo_entry.delete(0, tk.END)
            self.photo_entry.configure(state='readonly')
            
            # Set home location
            home_coord = user['home_location']
            for city, coord in CITIES.items():
                if coord == home_coord:
                    self.home_city.set(city)
                    break
            
            # Set office location
            office_coord = user['office_location']
            for office, coord in OFFICE_LOCATIONS.items():
                if coord == office_coord:
                    self.office_loc.set(office)
                    break
                    
            self.hide_loading()
            
            # Scroll to make form visible
            self.window.update()
            self.name_entry.focus()
            
        except Exception as e:
            print(f"Error starting edit: {str(e)}")
            self.hide_loading()
            messagebox.showerror(
                "Error",
                f"Gagal memulai edit: {str(e)}",
                parent=self.window
            )

    def delete_user(self):
        """Delete selected user with improved confirmation"""
        try:
            # Get selected item
            selection = self.user_tree.selection()
            if not selection:
                messagebox.showwarning(
                    "Peringatan",
                    "Silakan pilih user yang akan dihapus!",
                    parent=self.window
                )
                return
                
            # Get user data
            item = self.user_tree.item(selection[0])
            user_id = item['values'][0]
            user_name = item['values'][1]
            photo_path = item['values'][2]
            
            # Show confirmation dialog
            confirm = messagebox.askyesno(
                "Konfirmasi Penghapusan",
                f"Anda yakin ingin menghapus user {user_name}?\n\n" + 
                "Semua data absensi user ini juga akan dihapus.",
                parent=self.window,
                icon='warning'
            )
            
            if not confirm:
                return
            
            # Show loading
            self.show_loading(f"Menghapus user {user_name}...")
                
            # Delete photo file
            try:
                photo_full_path = os.path.join("data", "user_faces", photo_path)
                if os.path.exists(photo_full_path):
                    os.remove(photo_full_path)
            except Exception as e:
                print(f"Error deleting photo: {str(e)}")
                
            # Delete from database
            if self.db.delete_user(user_id):
                self.hide_loading()
                messagebox.showinfo(
                    "Sukses",
                    f"User {user_name} berhasil dihapus!",
                    parent=self.window
                )
                self.load_users()
            else:
                raise Exception("Gagal menghapus user")
                
        except Exception as e:
            print(f"Error deleting user: {str(e)}")
            self.hide_loading()
            messagebox.showerror(
                "Error",
                f"Gagal menghapus user: {str(e)}",
                parent=self.window
            )

    def cancel_edit(self):
        """Cancel editing user"""
        self.current_user_id = None
        self.submit_button.configure(text="Tambah User")
        self.cancel_button.pack_forget()
        self.clear_form()

    def clear_form(self):
        """Clear all form fields"""
        self.name_entry.delete(0, tk.END)
        self.selected_photo_path = None
        self.photo_entry.configure(state='normal')
        self.photo_entry.delete(0, tk.END)
        self.photo_entry.configure(state='readonly')
        self.home_city.set('')
        self.office_loc.set('')
        self.name_entry.focus()

    def load_users(self):
        """Load users into treeview with improved visual feedback"""
        try:
            self.show_loading("Memuat daftar user...")
            
            # Clear existing items
            for item in self.user_tree.get_children():
                self.user_tree.delete(item)
                
            # Get users from database
            users = self.db.get_users()
            
            # Insert into treeview
            alternate = False  # For alternating row colors
            for user in users:
                self.user_tree.insert(
                    '',
                    'end',
                    values=(
                        user['id'],
                        user['name'],
                        user['photo_path'],
                        user['created_at']
                    ),
                    tags=('alternate',) if alternate else ()
                )
                alternate = not alternate
                
            self.hide_loading()
            
        except Exception as e:
            print(f"Error loading users: {str(e)}")
            self.hide_loading()
            messagebox.showerror(
                "Error",
                f"Gagal memuat daftar user: {str(e)}",
                parent=self.window
            )

    def load_attendance(self):
        """Load attendance records into treeview with improved visual feedback"""
        try:
            self.show_loading("Memuat data absensi...")
            
            # Clear existing items
            for item in self.attendance_tree.get_children():
                self.attendance_tree.delete(item)
                
            # Get attendance records
            records = self.db.get_attendance()
            
            # Insert into treeview with alternating colors
            alternate = False
            for record in records:
                # Style row based on status
                status = record['status']
                if status == "Present":
                    tag = 'present'
                elif status == "Late":
                    tag = 'late'
                else:
                    tag = 'absent'
                
                if alternate:
                    tag = f"{tag}_alternate"
                
                self.attendance_tree.insert(
                    '',
                    'end',
                    values=(
                        record['id'],
                        record['name'],
                        record['date'],
                        record['time_in'],
                        record['time_out'] or '-',
                        record['mode'],
                        record['status'],
                        record['location'] or '-'
                    ),
                    tags=(tag,)
                )
                alternate = not alternate
                
            self.hide_loading()
            
        except Exception as e:
            print(f"Error loading attendance: {str(e)}")
            self.hide_loading()
            messagebox.showerror(
                "Error",
                f"Gagal memuat data absensi: {str(e)}",
                parent=self.window
            )

    def process_photo(self, name):
        """Process and save photo file with size validation"""
        try:
            user_faces_dir = os.path.join("data", "user_faces")
            os.makedirs(user_faces_dir, exist_ok=True)
            
            # Create photo filename from user name
            base_name = name.lower().replace(' ', '_')
            photo_ext = os.path.splitext(self.selected_photo_path)[1].lower()
            photo_name = f"{base_name}{photo_ext}"
            photo_path = os.path.join(user_faces_dir, photo_name)
            
            # Handle duplicate filenames
            counter = 1
            while os.path.exists(photo_path):
                photo_name = f"{base_name}_{counter}{photo_ext}"
                photo_path = os.path.join(user_faces_dir, photo_name)
                counter += 1
            
            # Copy photo with feedback
            self.loading_label.config(text="Memproses foto...")
            self.window.update()
            
            shutil.copy2(self.selected_photo_path, photo_path)
            return photo_name
            
        except Exception as e:
            print(f"Error processing photo: {str(e)}")
            raise

    def export_attendance(self):
        """Export attendance records to Excel with progress feedback"""
        try:
            # Get file path
            filename = filedialog.asksaveasfilename(
                title="Export Laporan Absensi",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                parent=self.window
            )
            
            if not filename:
                return
                
            self.show_loading("Mengexport data absensi...")
                
            try:
                import openpyxl
            except ImportError:
                self.hide_loading()
                messagebox.showerror(
                    "Error",
                    "Modul openpyxl belum terinstall!\n" +
                    "Jalankan 'pip install openpyxl' di terminal.",
                    parent=self.window
                )
                return
                
            # Get attendance records
            records = self.db.get_attendance()
            
            # Create workbook and sheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Laporan Absensi"
            
            # Write headers with styling
            headers = ["ID", "Nama", "Tanggal", "Jam Masuk", "Jam Keluar", "Mode", "Status", "Lokasi"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
            
            # Write data
            for row, record in enumerate(records, 2):
                self.loading_label.config(text=f"Mengexport data {row-1} dari {len(records)}...")
                self.window.update()
                
                ws.cell(row=row, column=1, value=record['id'])
                ws.cell(row=row, column=2, value=record['name'])
                ws.cell(row=row, column=3, value=record['date'])
                ws.cell(row=row, column=4, value=record['time_in'])
                ws.cell(row=row, column=5, value=record['time_out'])
                ws.cell(row=row, column=6, value=record['mode'])
                ws.cell(row=row, column=7, value=record['status'])
                ws.cell(row=row, column=8, value=record['location'])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(column[0].column)
                
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(filename)
            
            self.hide_loading()
            messagebox.showinfo(
                "Sukses",
                "Laporan absensi berhasil di-export!",
                parent=self.window
            )
            
        except Exception as e:
            print(f"Error exporting attendance: {str(e)}")
            self.hide_loading()
            messagebox.showerror(
                "Error", 
                f"Gagal export laporan: {str(e)}",
                parent=self.window
            )        